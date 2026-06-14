import openpyxl, json, sys, re
sys.stdout.reconfigure(encoding='utf-8')

# ===========================
# 1. READ ALL ANALYSIS SHEETS
# ===========================
wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\ANALYSIS OF NARRATIVE PLANS.xlsx', data_only=True, read_only=True)

analysis_orgs = {}
for sheet_name in wb.sheetnames:
    if sheet_name in ('SUM N-SCORE', 'SUM NARRATIVE'):
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    org_id = ''
    org_name = ''
    analysis = {}

    for i, row in enumerate(rows, 1):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        col2 = str(row[2]).strip() if row[2] else ''

        if i == 2:
            if row[1]: org_id = str(row[1]).strip()
            if row[2]: org_name = str(row[2]).strip()
            continue

        # Summary row
        if col0 == 'SUMMARY':
            vals = [str(v).strip() for v in row[1:] if v is not None and str(v).strip()]
            if vals: analysis['SUMMARY'] = ' | '.join(vals)
            continue

        # Vision
        if col1 == 'VISION':
            analysis['VISION'] = col2
            continue

        # Ambitions
        if col0 == 'Ambitions' and col2:
            if 'AMBITIONS' not in analysis:
                analysis['AMBITIONS'] = col2
            continue

        # Enablers
        if col0 == 'Enablers' and col2:
            if 'ENABLERS' not in analysis:
                analysis['ENABLERS'] = col2
            continue

        # Delivery (short) - first occurrence
        if col1 == 'DELIVERY' and 'DELIVERY_SHORT' not in analysis:
            analysis['DELIVERY_SHORT'] = col2
            continue

    # Now find dimensions by scanning col1 for known dimension names
    dim_header_to_key = {
        'GEOGRAPHY': 'GEOGRAPHY',
        '10YP ALIGNMENT': '10YP_ALIGNMENT',
        '10YP DELIVERABILITY': '10YP_DELIVERABILITY',
        'TRADE-OFFS': 'TRADE-OFFS',
        'PREVENTION READINESS': 'PREVENTION_READINESS',
        'DIGITAL READINESS': 'DIGITAL_READINESS',
        'CARE CLOSE TO HOME READINESS': 'CARE_CLOSE_TO_HOME',
        'WORKFORCE READINESS - CULTURE / WAYS OF WORKING': 'WORKFORCE_READINESS',
        'EQUITY READINESS': 'EQUITY_READINESS',
        'INNOVATION/ TRANSFORMATION': 'INNOVATION_TRANSFORMATION',
        'ADAPTABILITY': 'ADAPTABILITY',
    }

    for i, row in enumerate(rows, 1):
        col1 = str(row[1]).strip() if row[1] else ''
        col2 = str(row[2]).strip() if row[2] else ''
        if col1 in dim_header_to_key and col2:
            key = dim_header_to_key[col1]
            if key not in analysis:
                analysis[key] = col2

    # OTHER items - between last dimension and DELIVERY
    in_other = False
    other_items = []
    for i, row in enumerate(rows, 1):
        col1 = str(row[1]).strip() if row[1] else ''
        col2 = str(row[2]).strip() if row[2] else ''
        col0 = str(row[0]).strip() if row[0] else ''

        if col1 == 'ADAPTABILITY':
            in_other = True
            continue
        if col1 == 'DELIVERY':
            break
        if in_other and col2:
            other_items.append(col2)

    if other_items:
        # Deduplicate while preserving order
        seen = set()
        unique = []
        for item in other_items:
            if item not in seen:
                seen.add(item)
                unique.append(item)
        analysis['OTHER'] = unique

    # DELIVERY table: find row where col1 is DELIVERY (second occurrence for some sheets)
    delivery_row = None
    delivery_count = 0
    for i, row in enumerate(rows, 1):
        col1 = str(row[1]).strip() if row[1] else ''
        col0 = str(row[0]).strip() if row[0] else ''
        if col1 == 'DELIVERY':
            delivery_count += 1
            # Use the LAST occurrence of DELIVERY
            delivery_row = i

    if delivery_row:
        delivery = []
        for i in range(delivery_row - 1, len(rows)):
            row = rows[i]
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals:
                delivery.append(vals)
        if len(delivery) > 1:  # At least header + one data row
            analysis['DELIVERY'] = delivery

    if org_name:
        sid = org_id
        if sid.startswith('NW_LSC'): region = 'LSC'
        elif sid.startswith('NW_CM'): region = 'C&M'
        elif sid.startswith('NW_GM'): region = 'GM'
        else: region = ''
        analysis_orgs[org_name] = {
            'id': org_id,
            'region': region,
            'analysis': analysis
        }

# ===========================
# 2. MANUAL ANALYSIS -> CONTACT MAPPINGS
# ===========================
analysis_to_contact = {
    'Alder Hey Children\'s NHSFT': 'Alder Hey Children\'s NHS Foundation Trust',
    'Blackpool Teaching NHSFT': 'Blackpool Teaching Hospital NHS Foundation Trust',
    'Bolton NHS Foundation Trust': 'Bolton NHS Foundation Trust',
    'Cheshire & Merseyside ICB': None,
    'Cheshire and Wirral Partnership NHS Foundation Trust': 'Cheshire and Wirral Partnership NHS Foundation Trust',
    'Clatterbridge Cancer Centre NHS Foundation Trust': 'Clatterbridge Cancer Centre NHS Foundation Trust',
    'Countess of Chester Hospital NHSFT': 'Countess of Chester Hospital NHS Foundation Trust',
    'East Cheshire NHS Trust': 'East Cheshire NHS Trust',
    'East Lancashire Hospitals NHST': 'East Lancashire Hospitals NHS Trust',
    'Greater Manchester ICB': None,
    'Greater Manchester Mental Health NHS Foundation Trust': 'Greater Manchester Mental Health NHS Foundation Trust',
    'Lancashire & South Cumbria ICB': None,
    'Lancashire Teaching Hospitals NHSFT': 'Lancashire Teaching Hospitals NHS Foundation Trust',
    'Lancashire and South Cumbria NHS Foundation Trust': 'Lancashire and South Cumbria NHS Foundation Trust',
    'Manchester University NHS Foundation Trust': 'Manchester University NHS Foundation Trust',
    'Mersey Care NHS Foundation Trust': 'Mersey Care NHS Foundation Trust',
    'Mersey and West Lancashire Teaching Hospitals NHS Trust': 'Mersey and West Lancashire Teaching Hospitals NHS Trust',
    'Mid Cheshire Hospital NHS Foundation Trust': 'Mid Cheshire Hospitals NHS Foundation Trust',
    'North Cheshire and Mersey NHSFT - Warrington and Halton Teaching Hospitals NHSFT (WHH) and Bridgewater Community Healthcare NHSFT (BCH) will become a single organisation from April 2026': 'Warrington and Halton Hospitals NHS Foundation Trust',
    'North West Ambulance Service NHS Trust': 'North West Ambulance Service NHS Trust',
    'Northern Care Alliance NHS Foundation Trust': 'Northern Care Alliance NHS Group',
    'Pennine Care NHS Foundation Trust': 'Pennine Care NHS Foundation Trust',
    'Stockport NHS Foundation Trust (SFT) & Tameside & Glossop Integrated Care NHS Foundation Trust (TGICFT)': 'Stockport NHS Foundation Trust',
    'Tameside & Glossop Integrated Care NHS Foundation Trust (TGICFT)': 'Tameside & Glossop Integrated Care NHS Foundation Trust',
    'The Christie NHS Foundation Trust': 'The Christie NHS Foundation Trust',
    'The Walton Centre NHSFT': 'The Walton Centre NHS Foundation Trust',
    'University Hospitals of Liverpool Group (UHLG) currently comprised of Liverpool University Hospital (LUHFT), Liverpool Women\u2019s (LWH) and Liverpool Heart and Chest (LHCH)': None,
    'University Hospitals of Morecambe Bay NHSFT': 'University Hospitals of Morecambe Bay NHS Foundation Trust',
    'Wirral Community Health and Care NHS FT\nand\nWirral University Teaching Hospital NHS FT': 'Wirral Community Health & Care NHS Foundation Trust',
    'Wrightington, Wigan and Leigh Teaching Hospitals NHS Foundation Trust': 'Wrightington, Wigan and Leigh NHS Foundation Trust',
}

# ===========================
# 3. READ ALL CONTACTS
# ===========================
wb2 = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\Strategy Contacts - upadted 5.2.26.xlsx', data_only=True, read_only=True)

org_name_canonical = {
    'Mersey Care NHS foundation Trust': 'Mersey Care NHS Foundation Trust',
    'Northern Care Alliance NHS Group ': 'Northern Care Alliance NHS Group',
    'East Cheshire': 'East Cheshire NHS Trust',
    'East Lancs': 'East Lancashire Hospitals NHS Trust',
    'Wirral University Teaching Hospital NHS Foundation Trust ': 'Wirral University Teaching Hospital NHS Foundation Trust',
    'Bridgewater Community Healthcare NHS Foundation Trust': 'Bridgewater Community Healthcare NHS Foundation Trust',
    'Stockport NHS Foundation Trust  & Tameside & Glossop Integrated Care NHS Foundation Trust': 'Stockport NHS Foundation Trust',
}
region_normalize = {'LANCS': 'LSC'}

all_contacts = []
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        org = str(row[2]).strip()
        org = org_name_canonical.get(org, org)
        region = str(row[0]).strip() if row[0] else ''
        region = region_normalize.get(region, region)
        all_contacts.append({
            'region': region,
            'type': str(row[1]).strip() if row[1] else '',
            'org': org,
            'job': str(row[3]).strip() if row[3] else '',
            'name': str(row[4]).strip() if row[4] else '',
            'email': str(row[5]).strip() if row[5] else '',
            'phone': str(row[6]).strip() if row[6] else '',
            'pa': str(row[7]).strip() if row[7] else '',
            'pa_phone': str(row[8]).strip() if row[8] else ''
        })

# ===========================
# 4. BUILD ORG MAP
# ===========================
org_map = {}

for c in all_contacts:
    org = c['org']
    if org not in org_map:
        org_map[org] = {
            'org': org,
            'region': c['region'],
            'type': c['type'],
            'contacts': []
        }
    else:
        if c['region'] and not org_map[org]['region']:
            org_map[org]['region'] = c['region']
        if not org_map[org]['type'] and c['type']:
            org_map[org]['type'] = c['type']
    dup = False
    for existing in org_map[org]['contacts']:
        if existing['name'] == c['name'] and existing['email'] == c['email'] and existing['job'] == c['job']:
            dup = True
            break
    if not dup and c['name']:
        org_map[org]['contacts'].append(c)

# Merge analysis
for a_name, a_data in sorted(analysis_orgs.items()):
    target = analysis_to_contact.get(a_name, None)
    if target is None:
        continue
    if target in org_map:
        org_map[target]['analysis'] = a_data['analysis']
        if not org_map[target]['region'] and a_data['region']:
            org_map[target]['region'] = a_data['region']
    else:
        org_map[a_name] = {
            'org': a_name,
            'region': a_data.get('region', ''),
            'type': '',
            'contacts': [],
            'analysis': a_data['analysis']
        }

# Special: University Hospitals of Liverpool Group
liverpool_group = analysis_orgs.get('University Hospitals of Liverpool Group (UHLG) currently comprised of Liverpool University Hospital (LUHFT), Liverpool Women\u2019s (LWH) and Liverpool Heart and Chest (LHCH)', None)
if liverpool_group:
    for liv_org in ['Liverpool University Hospitals NHS Foundation Trust', 'Liverpool Women\'s Hospital NHS Foundation Trust', 'Liverpool Heart and Chest Hospital NHS Foundation Trust']:
        if liv_org in org_map:
            org_map[liv_org]['analysis'] = liverpool_group['analysis']
            if not org_map[liv_org]['region']:
                org_map[liv_org]['region'] = 'C&M'

# ===========================
# 5. VERIFY: print analysis dimensions per org
# ===========================
print("=== Analysis verification ===")
for k, v in sorted(org_map.items()):
    if v.get('analysis'):
        a = v['analysis']
        dims_present = []
        for d in ['GEOGRAPHY', '10YP_ALIGNMENT', '10YP_DELIVERABILITY', 'TRADE-OFFS', 'PREVENTION_READINESS', 'DIGITAL_READINESS', 'CARE_CLOSE_TO_HOME', 'WORKFORCE_READINESS', 'EQUITY_READINESS', 'INNOVATION_TRANSFORMATION', 'ADAPTABILITY']:
            if a.get(d):
                dims_present.append(d)
        print(f"  {k[:50]:50s} | dims={len(dims_present):2d} | other={len(a.get('OTHER',[]))} | delivery={'Y' if a.get('DELIVERY') else 'N'}")

# ===========================
# 6. GENERATE HTML
# ===========================
html = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NHS Organisation Database</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;background:#f0f4f8;color:#1a1a2e}
.header{background:#005eb8;color:white;padding:24px 32px 20px;position:sticky;top:0;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,.15)}
.header h1{font-size:22px;font-weight:600;margin-bottom:12px}
.search-wrap{display:flex;gap:10px;max-width:700px}
.search-wrap input{flex:1;padding:10px 14px;border:2px solid transparent;border-radius:6px;font-size:15px;outline:none;transition:.2s}
.search-wrap input:focus{border-color:#ffb81c}
.search-wrap select{padding:10px 12px;border-radius:6px;border:2px solid transparent;font-size:14px;background:white;cursor:pointer}
.stats{font-size:13px;margin-top:8px;opacity:.85}
.container{max-width:1400px;margin:0 auto;padding:20px}
.org-card{background:white;border-radius:10px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden;transition:.2s}
.org-card:hover{box-shadow:0 3px 12px rgba(0,0,0,.12)}
.org-header{padding:16px 20px;cursor:pointer;display:flex;justify-content:space-between;align-items:flex-start;background:#fafcff;border-bottom:1px solid #e8edf2;user-select:none}
.org-header h2{font-size:16px;font-weight:600;color:#005eb8;margin-bottom:4px}
.org-header .arrow{font-size:12px;color:#6b7a8a;transition:.2s;margin-top:4px}
.org-header .arrow.open{transform:rotate(180deg)}
.org-body{display:none;padding:0}
.org-body.open{display:block}
.analysis-section{padding:16px 20px;border-bottom:1px solid #eef2f7}
.analysis-section:last-child{border-bottom:none}
.analysis-section h3{font-size:14px;font-weight:600;color:#1a1a2e;margin-bottom:8px;padding-bottom:6px;border-bottom:2px solid #005eb8;display:inline-block}
.analysis-section p,.analysis-section div.txt{font-size:13px;line-height:1.6;color:#2d3748;margin-bottom:6px;white-space:pre-wrap}
.analysis-section p:last-child{margin-bottom:0}
.delivery-table{width:100%;border-collapse:collapse;font-size:13px;margin-top:8px}
.delivery-table th{background:#eef2f7;padding:8px 10px;text-align:left;font-weight:600}
.delivery-table td{padding:8px 10px;border-bottom:1px solid #e8edf2;vertical-align:top}
.contacts-table{width:100%;border-collapse:collapse;font-size:13px}
.contacts-table th{background:#eef2f7;padding:8px 10px;text-align:left;font-weight:600;position:sticky;top:84px}
.contacts-table td{padding:8px 10px;border-bottom:1px solid #eef2f7;vertical-align:top}
.contacts-table tr:hover{background:#f8faff}
.contacts-toggle{padding:12px 20px;background:#f8faff;cursor:pointer;display:flex;align-items:center;gap:8px;font-size:13px;font-weight:500;color:#005eb8;border-top:1px solid #e8edf2;user-select:none}
.contacts-toggle:hover{background:#eef4ff}
.no-results{text-align:center;padding:60px 20px;color:#6b7a8a;font-size:16px}
.tag{display:inline-block;font-size:11px;padding:2px 8px;border-radius:4px;margin-right:4px;font-weight:500;margin-bottom:2px}
.tag-LSC{background:#e8f4e8;color:#2d7d2d}
.tag-C\\&M{background:#e8eeff;color:#2d4d9e}
.tag-GM{background:#fff3e0;color:#b85c00}
.tag-job{background:#f0f0f5;color:#4a4a6a}
.tag-analysis{background:#e6f7ed;color:#1a7a3a;font-weight:600}
.email-link{color:#005eb8;text-decoration:none}
.email-link:hover{text-decoration:underline}
.mobile-label{display:none;font-weight:600;color:#4a5568}
@media(max-width:768px){
  .header{padding:16px}
  .header h1{font-size:18px}
  .search-wrap{flex-direction:column}
  .org-header{padding:12px 16px}
  .org-header h2{font-size:14px}
  .contacts-table th{display:none}
  .contacts-table td{display:block;padding:4px 10px;border-bottom:none}
  .contacts-table tr{border-bottom:1px solid #eef2f7;padding:6px 0}
  .mobile-label{display:inline-block;width:90px}
}
</style>
</head>
<body>
<div class="header">
  <h1>NHS Organisation Database</h1>
  <div class="search-wrap">
    <input type="text" id="search" placeholder="Search organisations, contacts, job titles..." oninput="filter()">
    <select id="regionFilter" onchange="filter()">
      <option value="">All Regions</option>
    </select>
  </div>
  <div class="stats" id="stats"></div>
</div>
<div class="container" id="container"></div>
<script>
var ORGS = ''' + json.dumps(org_map, ensure_ascii=False) + ''';

var regions = new Set();
Object.values(ORGS).forEach(function(o) { if(o.region) regions.add(o.region); });
var sel = document.getElementById('regionFilter');
Array.from(regions).sort().forEach(function(r) {
  var opt = document.createElement('option');
  opt.value = r; opt.textContent = r; sel.appendChild(opt);
});

function esc(s) {
  if(!s) return '';
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/\\xa0/g,' ');
}

function formatTxt(t) {
  return esc(t).replace(/\\n/g, '<br>');
}

function render() {
  var q = document.getElementById('search').value.toLowerCase().trim();
  var reg = document.getElementById('regionFilter').value;
  var container = document.getElementById('container');
  var html = '';
  var count = 0;
  var keys = Object.keys(ORGS).sort();

  for(var ki = 0; ki < keys.length; ki++) {
    var k = keys[ki];
    var o = ORGS[k];
    if(reg && o.region !== reg) continue;
    if(q) {
      var match = k.toLowerCase().includes(q);
      if(!match) {
        for(var ci = 0; ci < o.contacts.length; ci++) {
          var c = o.contacts[ci];
          if(c.name.toLowerCase().includes(q) || c.job.toLowerCase().includes(q) || c.email.toLowerCase().includes(q) || c.phone.toLowerCase().includes(q)) {
            match = true; break;
          }
        }
      }
      if(!match) continue;
    }
    count++;

    var analysisContent = '';
    if(o.analysis) {
      var a = o.analysis;
      var secs = [];
      if(a.SUMMARY) secs.push({l:'Summary',t:a.SUMMARY});
      if(a.VISION) secs.push({l:'Vision',t:a.VISION});
      if(a.AMBITIONS) secs.push({l:'Ambitions',t:a.AMBITIONS});
      if(a.ENABLERS) secs.push({l:'Enablers',t:a.ENABLERS});
      if(a.DELIVERY_SHORT) secs.push({l:'Delivery Overview',t:a.DELIVERY_SHORT});
      var dims = ['GEOGRAPHY','10YP_ALIGNMENT','10YP_DELIVERABILITY','TRADE-OFFS','PREVENTION_READINESS','DIGITAL_READINESS','CARE_CLOSE_TO_HOME','WORKFORCE_READINESS','EQUITY_READINESS','INNOVATION_TRANSFORMATION','ADAPTABILITY'];
      var lbls = ['Geography','10YP Alignment','10YP Deliverability','Trade-offs','Prevention Readiness','Digital Readiness','Care Close to Home','Workforce Readiness','Equity Readiness','Innovation / Transformation','Adaptability'];
      for(var di = 0; di < dims.length; di++) {
        if(a[dims[di]]) secs.push({l:lbls[di],t:a[dims[di]]});
      }
      if(a.OTHER) {
        var ot = Array.isArray(a.OTHER) ? a.OTHER.join('\\n') : a.OTHER;
        secs.push({l:'Other',t:ot});
      }
      for(var si = 0; si < secs.length; si++) {
        analysisContent += '<div class="analysis-section"><h3>' + esc(secs[si].l) + '</h3><div class="txt">' + formatTxt(secs[si].t) + '</div></div>';
      }
      if(a.DELIVERY) {
        analysisContent += '<div class="analysis-section"><h3>Delivery</h3><table class="delivery-table"><thead><tr>';
        for(var hi = 0; hi < a.DELIVERY[0].length; hi++) {
          analysisContent += '<th>' + esc(a.DELIVERY[0][hi]) + '</th>';
        }
        analysisContent += '</tr></thead><tbody>';
        for(var dr = 1; dr < a.DELIVERY.length; dr++) {
          analysisContent += '<tr>';
          for(var dci = 0; dci < a.DELIVERY[dr].length; dci++) {
            analysisContent += '<td>' + formatTxt(a.DELIVERY[dr][dci]) + '</td>';
          }
          analysisContent += '</tr>';
        }
        analysisContent += '</tbody></table></div>';
      }
    }

    var contactRows = '';
    for(var ci = 0; ci < o.contacts.length; ci++) {
      var c = o.contacts[ci];
      contactRows += '<tr>';
      contactRows += '<td><span class="mobile-label">Name:</span>' + esc(c.name) + '</td>';
      contactRows += '<td><span class="mobile-label">Job:</span><span class="tag tag-job">' + esc(c.job) + '</span></td>';
      contactRows += '<td><span class="mobile-label">Email:</span>' + (c.email ? '<a class="email-link" href="mailto:' + esc(c.email) + '">' + esc(c.email) + '</a>' : '') + '</td>';
      contactRows += '<td><span class="mobile-label">Phone:</span>' + esc(c.phone) + '</td>';
      contactRows += '<td><span class="mobile-label">PA:</span>' + esc(c.pa) + '</td>';
      contactRows += '<td><span class="mobile-label">PA Phone:</span>' + esc(c.pa_phone) + '</td>';
      contactRows += '</tr>';
    }

    var rtc = 'tag-' + (o.region || '').replace(/&/g,'').replace(/ /g,'');
    html += '<div class="org-card">';
    html += '<div class="org-header" onclick="toggleOrg(this)">';
    html += '<div><h2>' + esc(k) + '</h2>';
    html += '<span class="tag ' + rtc + '">' + esc(o.region) + '</span> ';
    html += '<span class="tag tag-job">' + esc(o.type) + '</span> ';
    html += '<span style="font-size:12px;color:#6b7a8a">' + o.contacts.length + ' contact' + (o.contacts.length !== 1 ? 's' : '') + '</span>';
    if(o.analysis) html += ' <span class="tag tag-analysis">Analysis</span>';
    html += '</div>';
    html += '<span class="arrow">&#9660;</span></div>';
    html += '<div class="org-body">';
    if(o.analysis) html += analysisContent;
    if(contactRows) {
      html += '<div class="contacts-toggle" onclick="toggleContacts(this)">';
      html += '<span>&#9776;</span> Contacts (' + o.contacts.length + ') <span class="arrow" style="margin-left:auto">&#9660;</span>';
      html += '</div>';
      html += '<div style="overflow-x:auto;display:none"><table class="contacts-table"><thead><tr>';
      html += '<th>Name</th><th>Job Title</th><th>Email</th><th>Phone</th><th>PA</th><th>PA Phone</th>';
      html += '</tr></thead><tbody>' + contactRows + '</tbody></table></div>';
    }
    html += '</div></div>';
  }

  if(count === 0) {
    html = '<div class="no-results"><p style="font-size:48px;margin-bottom:12px">&#128270;</p><p>No organisations match your search</p></div>';
  }
  container.innerHTML = html;
  document.getElementById('stats').textContent = 'Showing ' + count + ' of ' + keys.length + ' organisations';
}

function toggleOrg(el) {
  var body = el.nextElementSibling;
  var arrow = el.querySelector('.arrow');
  if(body) body.classList.toggle('open');
  if(arrow) arrow.classList.toggle('open');
}

function toggleContacts(el) {
  var table = el.nextElementSibling;
  var arrow = el.querySelector('.arrow');
  if(!table) return;
  if(table.style.display === 'none') { table.style.display = 'block'; if(arrow) arrow.classList.add('open'); }
  else { table.style.display = 'none'; if(arrow) arrow.classList.remove('open'); }
}

function filter() { render(); }
document.addEventListener('DOMContentLoaded', filter);
</script>
</body>
</html>'''

with open(r'C:\Users\Administrator\Downloads\NHS Big List\database.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f'\nTotal orgs in DB: {len(org_map)}')
print(f'Total contacts: {len(all_contacts)}')
