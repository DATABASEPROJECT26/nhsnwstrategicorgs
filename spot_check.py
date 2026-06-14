import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')

wb2 = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\Strategy Contacts - upadted 5.2.26.xlsx', data_only=True, read_only=True)

# Spot-check a few specific contacts from the source
print("=== SPOT CHECK: Key contacts from source ===")
checked = []

# Check contacts from sheet 2 (Trust Contacts) which has different org names
ws2 = wb2['Trust Contacts']
for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=30, values_only=True), 2):
    if row[2] and row[4]:
        print(f"Row {i}: {str(row[4]).strip():30s} | {str(row[2]).strip():50s} | {str(row[3]).strip() if row[3] else ''}")

print()
print("=== Contacts with empty org field ===")
for name in wb2.sheetnames:
    ws = wb2[name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if (not row[2] or not str(row[2]).strip()) and row[4] and str(row[4]).strip():
            print(f"[{name}] Row {i}: {str(row[4]).strip()}")
            break  # just show first one

print()
print("=== Orgs that appear ONLY in Sheet 2 ===")
ws1 = wb2['North West Trust Contacts']
sheet1_orgs = set()
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[2]: sheet1_orgs.add(str(row[2]).strip().lower())

ws2 = wb2['Trust Contacts']
sheet2_orgs = set()
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[2]: sheet2_orgs.add(str(row[2]).strip().lower())

only_in_s2 = sheet2_orgs - sheet1_orgs
print(f"Orgs unique to Sheet 2: {only_in_s2}")
