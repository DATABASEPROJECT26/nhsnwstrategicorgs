import openpyxl
wb2 = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\Strategy Contacts - upadted 5.2.26.xlsx', data_only=True, read_only=True)

ws2 = wb2['Trust Contacts']
print("=== Trust Contacts sheet (first 35 rows) ===")
for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=35, values_only=True), 2):
    if row[2] and row[4]:
        org = str(row[2]).strip()[:50]
        name = str(row[4]).strip()[:25]
        job = str(row[3]).strip()[:30] if row[3] else ''
        region = str(row[0]).strip() if row[0] else ''
        print(f'Row {i}: [{region}] {name:25s} | {org:50s} | {job}')

print()
print("=== Orgs unique to Sheet 2 ===")
ws1 = wb2['North West Trust Contacts']
s1 = set()
for r in ws1.iter_rows(min_row=2, values_only=True):
    if r[2]: s1.add(str(r[2]).strip().lower())

s2 = set()
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[2]: s2.add(str(r[2]).strip().lower())

only2 = s2 - s1
for o in sorted(only2):
    print(f'  {o}')
