import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\ANALYSIS OF NARRATIVE PLANS.xlsx', data_only=True, read_only=True)

# Find where dimensions actually are in each sheet
# Dimension names as they appear in col1 of each sheet
dim_names = ['GEOGRAPHY', '10YP ALIGNMENT', '10YP DELIVERABILITY', 'TRADE-OFFS',
             'PREVENTION READINESS', 'DIGITAL READINESS', 'CARE CLOSE TO HOME READINESS',
             'WORKFORCE READINESS - CULTURE / WAYS OF WORKING', 'EQUITY READINESS',
             'INNOVATION/ TRANSFORMATION', 'ADAPTABILITY']

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUM N-SCORE', 'SUM NARRATIVE'):
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    print(f'\n=== {sheet_name} ===')
    org_name = str(rows[1][2]).strip() if rows[1][2] else ''
    print(f'  Org: {org_name}')
    
    # Check where VISION, Ambitions, Enablers are
    for i, r in enumerate(rows):
        if r[1] and str(r[1]).strip() in ['VISION']:
            print(f'  VISION at row {i+1}')
        if r[0] and str(r[0]).strip() in ['Ambitions']:
            print(f'  Ambitions at row {i+1}')
        if r[0] and str(r[0]).strip() in ['Enablers']:
            print(f'  Enablers at row {i+1}')
        if r[1] and str(r[1]).strip() in dim_names:
            print(f'  DIM "{str(r[1]).strip()}" at row {i+1}')
        if r[1] and str(r[1]).strip().upper() == 'DELIVERY':
            print(f'  DELIVERY at row {i+1}')
            break  # stop at delivery
