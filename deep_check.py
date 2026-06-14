import openpyxl, json, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

with open(r'C:\Users\Administrator\Downloads\NHS Big List\database.html', 'r', encoding='utf-8') as f:
    html = f.read()
start = html.find('var ORGS = ') + len('var ORGS = ')
depth = 0
end = start
for i in range(start, len(html)):
    if html[i] == '{': depth += 1
    elif html[i] == '}': depth -= 1
    if depth == 0: end = i + 1; break
html_orgs = json.loads(html[start:end])

canon_map = {
    'Mersey Care NHS foundation Trust': 'Mersey Care NHS Foundation Trust',
    'Northern Care Alliance NHS Group ': 'Northern Care Alliance NHS Group',
    'East Cheshire': 'East Cheshire NHS Trust',
    'East Lancs': 'East Lancashire Hospitals NHS Trust',
    'Wirral University Teaching Hospital NHS Foundation Trust ': 'Wirral University Teaching Hospital NHS Foundation Trust',
    'Stockport NHS Foundation Trust  & Tameside & Glossop Integrated Care NHS Foundation Trust': 'Stockport NHS Foundation Trust',
}

# Read source contacts ONCE and collect stats
wb2 = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\Strategy Contacts - upadted 5.2.26.xlsx', data_only=True, read_only=True)

total_rows = 0
no_name = 0
empty_org = 0
per_sheet = {}
source_contacts = []

for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if not row[2] or not str(row[2]).strip():
            empty_org += 1
            continue
        org = str(row[2]).strip()
        org = canon_map.get(org, org)
        if not row[4] or not str(row[4]).strip():
            no_name += 1
            continue
        count += 1
        name = str(row[4]).strip()
        email = str(row[5]).strip() if row[5] else ''
        job = str(row[3]).strip() if row[3] else ''
        source_contacts.append({'org': org, 'name': name, 'email': email, 'job': job})
    per_sheet[sheet_name] = count

print(f"Total rows (excluding header): {total_rows}")
print(f"Empty org: {empty_org}")
print(f"No name: {no_name}")
print(f"Contacts with name+org: {sum(per_sheet.values())}")
for s, c in per_sheet.items():
    print(f"  {s}: {c}")

html_total = sum(len(v['contacts']) for v in html_orgs.values())
print(f"\nHTML contacts: {html_total}")
print(f"Difference: {sum(per_sheet.values()) - html_total}")

# Check for duplicate contacts in source
pairs = Counter()
for sc in source_contacts:
    pairs[(sc['org'], sc['name'], sc['email'])] += 1
multiples = {k: v for k, v in pairs.items() if v > 1}
print(f"\nDuplicate contacts (same org+name+email in source): {len(multiples)}")
for (org, name, email), count in sorted(multiples.items(), key=lambda x: -x[1])[:15]:
    print(f"  {name} @ {org} ({email}) appears {count}x")

# Build HTML lookup by (name, email)
html_lookup = {}
for org_name, org_data in html_orgs.items():
    for hc in org_data['contacts']:
        html_lookup[(hc['name'], hc['email'])] = True

# Find contacts in source but not HTML
missing = 0
for sc in source_contacts:
    if (sc['name'], sc['email']) not in html_lookup:
        missing += 1
        if missing <= 25:
            print(f"  MISSING: {sc['name']} ({sc['org']}) - {sc['job']} [{sc['email']}]")

print(f"\nTotal missing from HTML: {missing}")
