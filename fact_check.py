import openpyxl, json, sys, re

sys.stdout.reconfigure(encoding='utf-8')

# Load HTML data
with open(r'C:\Users\Administrator\Downloads\NHS Big List\database.html', 'r', encoding='utf-8') as f:
    html = f.read()

# Extract JSON from HTML - find var ORGS =  then match braces
start = html.find('var ORGS = ') + len('var ORGS = ')
depth = 0
end = start
for i in range(start, len(html)):
    if html[i] == '{': depth += 1
    elif html[i] == '}': depth -= 1
    if depth == 0:
        end = i + 1
        break
html_orgs = json.loads(html[start:end])

print("=" * 80)
print("CHECK 1: CONTACTS ACCURACY")
print("=" * 80)

# Read contacts source
wb2 = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\Strategy Contacts - upadted 5.2.26.xlsx', data_only=True, read_only=True)

source_contacts = []
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row[2]:
            source_contacts.append({
                'sheet': sheet_name,
                'row': i,
                'region': str(row[0]).strip() if row[0] else '',
                'type': str(row[1]).strip() if row[1] else '',
                'org': str(row[2]).strip(),
                'job': str(row[3]).strip() if row[3] else '',
                'name': str(row[4]).strip() if row[4] else '',
                'email': str(row[5]).strip() if row[5] else '',
                'phone': str(row[6]).strip() if row[6] else '',
                'pa': str(row[7]).strip() if row[7] else '',
                'pa_phone': str(row[8]).strip() if row[8] else ''
            })

# Count total html contacts
html_contacts = []
for org_name, org_data in html_orgs.items():
    for c in org_data.get('contacts', []):
        html_contacts.append(c)

print(f"Source contacts: {len(source_contacts)}")
print(f"HTML contacts:   {len(html_contacts)}")
print()

# Check for missing contacts: for each source contact, check if it exists in HTML
missing = 0
name_mismatches = 0
for sc in source_contacts:
    found = False
    for hc in html_contacts:
        if hc['name'] == sc['name'] and hc['email'] == sc['email'] and hc['org'] in html_orgs:
            # Check if the HTML org contains this contact
            for org_name, org_data in html_orgs.items():
                for c in org_data['contacts']:
                    if c['name'] == sc['name'] and c['email'] == sc['email']:
                        found = True
                        break
                if found: break
        if found: break
    if not found and sc['name']:
        missing += 1
        if missing <= 10:
            print(f"MISSING: Row {sc['row']} in [{sc['sheet']}]: {sc['name']} ({sc['org']})")

print(f"\nMissing contacts (in source but not HTML): {missing}")

# Check contacts in HTML but not source (should be none, but check for dupes)
html_only = 0
for hc in html_contacts:
    found_source = False
    for sc in source_contacts:
        if hc['name'] == sc['name'] and hc['email'] == sc['email'] and hc['job'] == sc['job']:
            found_source = True
            break
    if not found_source:
        html_only += 1

print(f"HTML-only contacts (possible dupes): {html_only}")

print()
print("=" * 80)
print("CHECK 2: ANALYSIS DATA ACCURACY")
print("=" * 80)

# Read analysis source
wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\ANALYSIS OF NARRATIVE PLANS.xlsx', data_only=True, read_only=True)

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUM N-SCORE', 'SUM NARRATIVE'):
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    source_org_name = ''
    if len(rows) >= 2 and rows[1][2]:
        source_org_name = str(rows[1][2]).strip()
    
    # Get summary from source row 7
    source_summary = ''
    if len(rows) >= 7 and rows[6]:
        vals = [str(v).strip() for v in rows[6][1:] if v is not None and str(v).strip()]
        if vals:
            source_summary = ' | '.join(vals)
    
    # Find this org in HTML
    found_in_html = False
    html_summary = ''
    for html_org_name, html_org_data in html_orgs.items():
        if html_org_data.get('analysis'):
            # Check if analysis name matches
            a = html_org_data['analysis']
            if a.get('SUMMARY', '') == source_summary and source_summary:
                found_in_html = True
                break
    
    # Print orgs that are missing analysis
    if not found_in_html and source_org_name:
        # Check if this org was intentionally excluded (ICBs)
        if 'ICB' not in source_org_name:
            print(f"MISSING ANALYSIS IN HTML: {source_org_name} (sheet {sheet_name})")

print()
print("=" * 80)
print("CHECK 3: ORG NAME MATCHING & REGIONS")
print("=" * 80)

# Check each analysis org matches correctly
analysis_to_html = {}
for html_org_name, html_org_data in html_orgs.items():
    if html_org_data.get('analysis'):
        a = html_org_data['analysis']
        if a.get('SUMMARY'):
            analysis_to_html[a['SUMMARY'][:50]] = html_org_name

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\NHS Big List\ANALYSIS OF NARRATIVE PLANS.xlsx', data_only=True, read_only=True)
for sheet_name in wb.sheetnames:
    if sheet_name in ('SUM N-SCORE', 'SUM NARRATIVE'):
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    source_org_name = ''
    if len(rows) >= 2 and rows[1][2]:
        source_org_name = str(rows[1][2]).strip()
    
    # Expected region from sheet prefix
    if sheet_name.startswith('NW_LSC'): expected_region = 'LSC'
    elif sheet_name.startswith('NW_CM'): expected_region = 'C&M'
    elif sheet_name.startswith('NW_GM'): expected_region = 'GM'
    else: expected_region = '?'
    
    # Find where analysis went
    found = False
    for html_org_name, html_org_data in html_orgs.items():
        if html_org_data.get('analysis'):
            # Check if summaries match (trimmed)
            a = html_org_data['analysis']
            source_sum = ''
            if len(rows) >= 7 and rows[6]:
                vals = [str(v).strip() for v in rows[6][1:] if v is not None and str(v).strip()]
                if vals:
                    source_sum = ' | '.join(vals)[:60]
            html_sum = (a.get('SUMMARY', '') or '')[:60]
            if source_sum and source_sum == html_sum:
                actual_region = html_org_data.get('region', '')
                region_match = 'OK' if actual_region == expected_region else f'MISMATCH (expected {expected_region}, got {actual_region})'
                print(f"[{region_match}] Sheet {sheet_name}: {source_org_name} -> {html_org_name}")
                found = True
                break
    if not found:
        if 'ICB' not in source_org_name:
            # Try to find where it went
            for html_org_name, html_org_data in html_orgs.items():
                if html_org_data.get('analysis') and source_org_name.lower().replace('nhsft','').replace('nhs','').strip() in html_org_name.lower():
                    actual_region = html_org_data.get('region', '')
                    region_match = 'OK' if actual_region == expected_region else f'MISMATCH'
                    print(f"[{region_match} - fuzzy] Sheet {sheet_name}: {source_org_name} -> {html_org_name}")
                    found = True
                    break
        if not found:
            print(f"[NOT FOUND] Sheet {sheet_name}: {source_org_name} (was excluded or missing)")

print()
print("=" * 80)
print("CHECK 4: DUPLICATE ORGANISATIONS IN HTML")
print("=" * 80)

# Check for similar org names
import difflib
names = list(html_orgs.keys())
for i in range(len(names)):
    for j in range(i+1, len(names)):
        ratio = difflib.SequenceMatcher(None, names[i].lower(), names[j].lower()).ratio()
        if ratio > 0.6:
            print(f"  SIMILAR: '{names[i]}' <-> '{names[j]}' (ratio: {ratio:.2f})")

print()
print("=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Source contacts: {len(source_contacts)}")
print(f"HTML contacts: {len(html_contacts)}")
print(f"Source analysis sheets: 30")
print(f"HTML orgs with analysis: {sum(1 for v in html_orgs.values() if v.get('analysis'))}")
print(f"Total orgs in HTML: {len(html_orgs)}")
